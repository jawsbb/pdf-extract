# -*- coding: utf-8 -*-
#!/usr/bin/env python3
"""
Interface Streamlit Pro - Extracteur Cadastral Français
Version client - Interface moderne et simplifiée
"""

import streamlit as st
import tempfile
import pandas as pd
from pathlib import Path
from pdf_extractor import PDFPropertyExtractor
import os
import io
import logging

# Configuration du logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Configuration de la page
st.set_page_config(
    page_title="Extracteur Cadastral Pro",
    page_icon="📋", 
    layout="wide",
    initial_sidebar_state="collapsed"
)

# CSS moderne
st.markdown("""
<style>
    .main-header {
        font-size: 3rem;
        font-weight: 700;
        background: linear-gradient(90deg, #1e3a8a 0%, #3b82f6 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        text-align: center;
        margin-bottom: 0.5rem;
    }
    .sub-header {
        font-size: 1.3rem;
        color: #64748b;
        text-align: center;
        margin-bottom: 2.5rem;
        font-weight: 400;
    }
    .upload-container {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        padding: 2.5rem;
        border-radius: 15px;
        margin: 2rem 0;
        text-align: center;
    }
    .upload-text {
        color: white;
        font-size: 1.2rem;
        font-weight: 500;
        margin-bottom: 1.5rem;
    }
    .results-container {
        background: linear-gradient(135deg, #f0f9ff 0%, #e0f2fe 100%);
        padding: 2rem;
        border-radius: 15px;
        border: 2px solid #0ea5e9;
        margin: 2rem 0;
    }
    .download-container {
        background: linear-gradient(135deg, #f0fdf4 0%, #dcfce7 100%);
        padding: 2rem;
        border-radius: 15px;
        border: 2px solid #22c55e;
        margin: 2rem 0;
    }
    .metric-box {
        background: white;
        padding: 1.5rem;
        border-radius: 12px;
        box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1);
        text-align: center;
        border: 1px solid #e2e8f0;
    }
    .metric-number {
        font-size: 2.5rem;
        font-weight: 700;
        color: #1e40af;
        margin-bottom: 0.5rem;
    }
    .metric-label {
        font-size: 0.9rem;
        color: #64748b;
        font-weight: 500;
    }
    .stButton > button {
        width: 100%;
        border-radius: 10px;
        font-weight: 600;
        font-size: 1.1rem;
        padding: 0.75rem;
        border: none;
        transition: all 0.3s ease;
    }
    .feature-badge {
        background: #f1f5f9;
        color: #475569;
        padding: 0.5rem 1rem;
        border-radius: 25px;
        font-size: 0.9rem;
        font-weight: 500;
        margin: 0.25rem;
        display: inline-block;
    }
    .footer {
        background: #f8f9fa;
        padding: 2rem;
        text-align: center;
        margin-top: 3rem;
        border-top: 1px solid #e9ecef;
        color: #6c757d;
        font-size: 0.9rem;
    }
    .footer-logo {
        font-weight: 600;
        color: #495057;
        font-size: 1.1rem;
    }
</style>
""", unsafe_allow_html=True)

def get_api_key():
    """Récupère la clé API OpenAI."""
    try:
        return st.secrets["OPENAI_API_KEY"]
    except:
        return os.getenv("OPENAI_API_KEY")

def initialize_extractor(temp_dir):
    """Initialise l'extracteur PDF."""
    api_key = get_api_key()
    if not api_key:
        st.error("🔑 Configuration API requise. Contactez l'administrateur.")
        st.stop()
    
    os.environ['OPENAI_API_KEY'] = api_key
    return PDFPropertyExtractor(
        input_dir=str(temp_dir / "input"),
        output_dir=str(temp_dir / "output")
    )

def create_excel_download(df, filename):
    """Crée un fichier Excel téléchargeable."""
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Données Cadastrales')
        
        # Formatter le fichier Excel
        workbook = writer.book
        worksheet = writer.sheets['Données Cadastrales']
        
        # Style pour les en-têtes (pour compatibilité)
        try:
            from openpyxl.styles import Font, PatternFill, Border, Side
            
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Appliquer le style aux en-têtes
            for col_num, value in enumerate(df.columns.values, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                worksheet.column_dimensions[cell.column_letter].width = 15
        except ImportError:
            # Fallback si openpyxl n'a pas les styles
            pass
    
    return output.getvalue()

def main():
    # En-tête moderne
    st.markdown('<h1 class="main-header">Extracteur Cadastral Pro</h1>', unsafe_allow_html=True)
    st.markdown('<p class="sub-header">Solution professionnelle d\'extraction de données cadastrales françaises</p>', unsafe_allow_html=True)
    
    # Badges des fonctionnalités
    st.markdown("""
    <div style="text-align: center; margin-bottom: 2rem;">
        <span class="feature-badge">IA Avancée</span>
        <span class="feature-badge">Export Excel & CSV</span>
        <span class="feature-badge">Traitement Rapide</span>
        <span class="feature-badge">Haute Précision</span>
    </div>
    """, unsafe_allow_html=True)
    
    # Section upload
    st.markdown("""
    <div class="upload-container">
        <div class="upload-text">
            Glissez-déposez vos fichiers PDF cadastraux ou cliquez pour les sélectionner
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    uploaded_files = st.file_uploader(
        "",
        type=['pdf'],
        accept_multiple_files=True,
        help="Formats supportés: PDF | Taille max: 200MB par fichier",
        label_visibility="collapsed"
    )
    
    # Options avancées dans un expander discret
    with st.expander("Options avancées", expanded=False):
        debug_mode = st.checkbox("Mode diagnostic (pour le support technique)", value=False)
        col1, col2 = st.columns(2)
        with col1:
            export_format = st.selectbox("Format d'export principal", ["Excel (.xlsx)", "CSV (.csv)"], index=0)
        with col2:
            include_empty = st.checkbox("Inclure les champs vides", value=True)
    
    if uploaded_files:
        st.success(f"{len(uploaded_files)} fichier(s) chargé(s) avec succès")
        
        # Informations sur les fichiers
        with st.expander("Détails des fichiers", expanded=False):
            for file in uploaded_files:
                file_size = len(file.getvalue()) / 1024 / 1024  # MB
                st.write(f"• **{file.name}** - {file_size:.1f} MB")
        
        if st.button("Démarrer l'extraction", type="primary", use_container_width=True):
            
            with st.status("Traitement en cours...", expanded=True) as status:
                
                # Créer un répertoire temporaire
                with tempfile.TemporaryDirectory() as temp_dir:
                    temp_path = Path(temp_dir)
                    input_dir = temp_path / "input"
                    output_dir = temp_path / "output"
                    input_dir.mkdir()
                    output_dir.mkdir()
                    
                    st.write("Préparation des fichiers...")
                    
                    # Sauvegarder les fichiers
                    saved_files = []
                    for uploaded_file in uploaded_files:
                        file_path = input_dir / uploaded_file.name
                        with open(file_path, "wb") as f:
                            f.write(uploaded_file.getbuffer())
                        saved_files.append(file_path)
                    
                    st.write("Initialisation du moteur IA...")
                    
                    try:
                        # Initialiser l'extracteur
                        extractor = initialize_extractor(temp_path)
                        
                        # Traitement
                        all_properties = []
                        
                        # Barre de progression
                        progress_bar = st.progress(0)
                        
                        for i, pdf_file in enumerate(saved_files):
                            st.write(f"Traitement: {pdf_file.name}")
                            progress_bar.progress((i + 1) / len(saved_files))
                            
                            if debug_mode:
                                with st.expander(f"Diagnostic - {pdf_file.name}"):
                                    try:
                                        structured = extractor.extract_tables_with_pdfplumber(pdf_file)
                                        st.json({
                                            "proprietes_baties": len(structured.get('prop_batie', [])),
                                            "proprietes_non_baties": len(structured.get('non_batie', [])),
                                            "sample_data": structured.get('non_batie', [{}])[:1]
                                        })
                                    except Exception as e:
                                        st.error(f"Erreur diagnostic: {e}")
                            
                            properties = extractor.process_like_make(pdf_file)
                            all_properties.extend(properties)
                        
                        status.update(label="Extraction terminée!", state="complete")
                    
                    except Exception as e:
                        st.error(f"Erreur lors du traitement: {e}")
                        all_properties = []
            
            # Affichage des résultats (en dehors du contexte status)
            if all_properties:
                # Préparation des données
                df = pd.DataFrame(all_properties)
                
                # Colonnes d'affichage
                display_columns = [
                    'department', 'commune', 'prefixe', 'section', 'numero', 
                    'contenance_ha', 'contenance_a', 'contenance_ca', 'droit_reel', 
                    'designation_parcelle', 'nom', 'prenom', 'numero_majic', 
                    'voie', 'post_code', 'city', 'id'
                ]
                
                df_display = df.reindex(columns=display_columns, fill_value='')
                
                # Conteneur des résultats
                st.markdown("""
                <div class="results-container">
                    <h3 style="color: #0ea5e9; margin-bottom: 1.5rem;">Résultats de l'extraction</h3>
                </div>
                """, unsafe_allow_html=True)
                
                # Métriques
                col1, col2, col3, col4 = st.columns(4)
                
                with col1:
                    st.markdown(f"""
                    <div class="metric-box">
                        <div class="metric-number">{len(all_properties)}</div>
                        <div class="metric-label">Propriétés</div>
                    </div>
                    """, unsafe_allow_html=True)
                
                with col2:
                    unique_owners = len(df[df['nom'].notna()]['nom'].unique())
                    st.markdown(f"""
                    <div class="metric-box">
                        <div class="metric-number">{unique_owners}</div>
                        <div class="metric-label">Propriétaires</div>
                    </div>
                    """, unsafe_allow_html=True)
                
                with col3:
                    files_processed = len(df['fichier_source'].unique())
                    st.markdown(f"""
                    <div class="metric-box">
                        <div class="metric-number">{files_processed}</div>
                        <div class="metric-label">Fichiers</div>
                    </div>
                    """, unsafe_allow_html=True)
                
                with col4:
                    completion_rate = (df['nom'].notna().sum() / len(df) * 100)
                    st.markdown(f"""
                    <div class="metric-box">
                        <div class="metric-number">{completion_rate:.0f}%</div>
                        <div class="metric-label">Complétude</div>
                    </div>
                    """, unsafe_allow_html=True)
                
                st.markdown("<br>", unsafe_allow_html=True)
                
                # Aperçu des données
                st.markdown("### Aperçu des données")
                st.dataframe(
                    df_display, 
                    use_container_width=True,
                    height=400
                )
                
                # Section téléchargement
                st.markdown("""
                <div class="download-container">
                    <h3 style="color: #22c55e; margin-bottom: 1.5rem;">Télécharger vos données</h3>
                </div>
                """, unsafe_allow_html=True)
                
                # Préparer les données pour export
                df_export = df_display.copy()
                
                # Renommer les colonnes pour l'export
                column_mapping = {
                    'department': 'Département',
                    'commune': 'Commune',
                    'prefixe': 'Préfixe',
                    'section': 'Section',
                    'numero': 'Numéro',
                    'contenance_ha': 'Contenance HA',
                    'contenance_a': 'Contenance A',
                    'contenance_ca': 'Contenance CA',
                    'droit_reel': 'Droit réel',
                    'designation_parcelle': 'Designation Parcelle',
                    'nom': 'Nom Propri',
                    'prenom': 'Prénom Propri',
                    'numero_majic': 'N°MAJIC',
                    'voie': 'Voie',
                    'post_code': 'CP',
                    'city': 'Ville',
                    'id': 'ID'
                }
                
                df_export = df_export.rename(columns=column_mapping)
                
                # Boutons de téléchargement
                col1, col2 = st.columns(2)
                
                with col1:
                    # Export Excel
                    excel_data = create_excel_download(df_export, "extraction_cadastrale.xlsx")
                    st.download_button(
                        label="Télécharger Excel",
                        data=excel_data,
                        file_name="extraction_cadastrale.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                
                with col2:
                    # Export CSV
                    csv_data = df_export.to_csv(index=False, sep=';', encoding='utf-8-sig')
                    st.download_button(
                        label="Télécharger CSV",
                        data=csv_data,
                        file_name="extraction_cadastrale.csv",
                        mime="text/csv",
                        use_container_width=True
                    )
            
            else:
                st.error("Aucune donnée n'a pu être extraite des fichiers PDF.")
                st.markdown("""
                **Suggestions:**
                - Vérifiez que les PDFs contiennent des tableaux cadastraux
                - Assurez-vous que les fichiers ne sont pas protégés
                - Contactez le support si le problème persiste
                """)
        
        # Informations techniques (en dehors de tous les contextes)
        if uploaded_files:
            with st.expander("Informations techniques", expanded=False):
                st.markdown("""
                **Format des données:**
                - **Excel**: Format .xlsx avec mise en forme automatique
                - **CSV**: Séparateur point-virgule (;), encodage UTF-8
                
                **Colonnes extraites:**
                - Informations parcellaires: Département, Commune, Préfixe, Section, Numéro
                - Contenance détaillée: HA (hectares), A (ares), CA (centiares)
                - Propriétaire: Nom, Prénom, N°MAJIC, Adresse complète
                - Identification unique: ID 14 caractères
                """)
    
    else:
        st.info("Sélectionnez un ou plusieurs fichiers PDF pour commencer l'extraction.")
        
        # Informations d'aide
        with st.expander("Comment utiliser cet outil", expanded=False):
            st.markdown("""
            ### Guide d'utilisation
            
            1. **Sélectionnez vos fichiers PDF** cadastraux français
            2. **Cliquez sur "Démarrer l'extraction"** pour lancer le traitement
            3. **Consultez les résultats** dans le tableau interactif
            4. **Téléchargez vos données** en Excel ou CSV
            
            ### Types de données extraites
            - **Informations parcellaires** (département, commune, section, numéro)
            - **Contenance détaillée** (hectares, ares, centiares)
            - **Propriétaires** (nom, prénom, adresse)
            - **Identifiants** (N°MAJIC, ID unique)
            
            ### Formats supportés
            - **PDF natifs** (générés numériquement)
            - **PDF scannés** (traitement OCR avancé)
            - **Multi-pages** (traitement automatique)
            """)
    
    # Footer avec copyright Level Up
    st.markdown("""
    <div class="footer">
        <div class="footer-logo">Level Up</div>
        <div>© 2024 Level Up. Tous droits réservés.</div>
        <div style="margin-top: 0.5rem; font-size: 0.8rem;">
            Solution d'extraction cadastrale professionnelle
        </div>
    </div>
    """, unsafe_allow_html=True)

if __name__ == "__main__":
    main() 